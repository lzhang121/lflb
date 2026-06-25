from openpyxl import load_workbook
from collections import Counter
import sys
sys.stdout.reconfigure(encoding="utf-8")

path = sys.argv[1]
wb = load_workbook(path)
print(wb.sheetnames)
for ws in wb.worksheets:
    print(ws.title, ws.max_row, ws.max_column)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
        print(repr(row))

texts = []
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip():
                texts.append(cell.value.strip())
print("count", len(texts))
for s, n in Counter(texts).most_common(120):
    print(repr(s), n)
