import argparse
import datetime as dt
import os
import time
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import urljoin

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://lflbssapp.com"
TENANT_COOKIE = "bssTenantCodeCookie"
DEFAULT_TENANT = "bssapp_esrdc"


class LoginFormParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.forms = []
        self._current_form = None
        self._current_textarea = None

    def handle_starttag(self, tag, attrs):
        attrs = dict(attrs)
        if tag == "form":
            self._current_form = {"attrs": attrs, "fields": {}}
        elif self._current_form and tag == "input":
            name = attrs.get("name")
            if name:
                self._current_form["fields"][name] = attrs.get("value", "")
        elif self._current_form and tag == "textarea":
            self._current_textarea = attrs.get("name")
            if self._current_textarea:
                self._current_form["fields"][self._current_textarea] = ""
        elif self._current_form and tag == "option":
            name = self._current_textarea
            if name and "value" in attrs and name not in self._current_form["fields"]:
                self._current_form["fields"][name] = attrs.get("value", "")

    def handle_endtag(self, tag):
        if tag == "form" and self._current_form:
            self.forms.append(self._current_form)
            self._current_form = None
        elif tag == "textarea":
            self._current_textarea = None

    def handle_data(self, data):
        if self._current_form and self._current_textarea:
            self._current_form["fields"][self._current_textarea] += data


def parse_login_form(html):
    parser = LoginFormParser()
    parser.feed(html)
    if not parser.forms:
        raise RuntimeError("No form found on the login page.")
    return parser.forms[0]


def fill_login_payload(payload, username, password, tenant_code):
    replacements = {
        "username": username,
        "userName": username,
        "loginName": username,
        "account": username,
        "password": password,
        "tenantCode": tenant_code,
        "tenant": tenant_code,
    }
    for key in list(payload.keys()):
        lower = key.lower()
        for needle, value in replacements.items():
            if needle.lower() == lower or needle.lower() in lower:
                payload[key] = value
                break
    return payload


def parse_args():
    parser = argparse.ArgumentParser(
        description="Export LFLB staff attendance data for the last N days to Excel."
    )
    parser.add_argument("--username", default=os.getenv("LFLB_USERNAME"))
    parser.add_argument("--password", default=os.getenv("LFLB_PASSWORD"))
    parser.add_argument("--tenant-code", default=os.getenv("LFLB_TENANT_CODE", DEFAULT_TENANT))
    parser.add_argument("--employees", default="employees.txt")
    parser.add_argument("--days", type=int, default=7)
    parser.add_argument("--end-date", default=dt.date.today().isoformat())
    parser.add_argument("--rows", type=int, default=50)
    parser.add_argument("--output", default=None)
    parser.add_argument("--debug", action="store_true")
    return parser.parse_args()


def load_employees(path):
    names = []
    for line in Path(path).read_text(encoding="utf-8-sig").splitlines():
        line = line.strip()
        if line and not line.startswith("#"):
            for chunk in split_employee_line(line):
                if "\t" in chunk:
                    name, code = [part.strip() for part in chunk.split("\t", 1)]
                elif "," in chunk:
                    name, code = [part.strip() for part in chunk.split(",", 1)]
                else:
                    name, code = chunk, ""
                if name:
                    names.append({"name": name, "code": code})
    if not names:
        raise ValueError(f"No employee names found in {path}")
    return names


def split_employee_line(line):
    normalized = (
        line.replace("、", ",")
        .replace("，", ",")
        .replace(";", ",")
        .replace("；", ",")
    )
    return [part.strip() for part in normalized.split(",") if part.strip()]


def date_range(end_date, days):
    end = dt.date.fromisoformat(end_date)
    start = end - dt.timedelta(days=days - 1)
    return [start + dt.timedelta(days=i) for i in range(days)]


def make_session(tenant_code):
    session = requests.Session()
    session.cookies.set(TENANT_COOKIE, tenant_code, domain="lflbssapp.com")
    session.headers.update(
        {
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-language": "ja,en-US;q=0.9,en;q=0.8",
            "origin": BASE_URL,
            "cache-control": "max-age=0",
            "user-agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/149.0.0.0 Safari/537.36"
            ),
            "sec-ch-ua": '"Google Chrome";v="149", "Chromium";v="149", "Not)A;Brand";v="24"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "same-origin",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
        }
    )
    return session


def login(session, username, password, tenant_code, debug=False):
    if not username or not password:
        raise ValueError("Missing username/password. Pass args or set LFLB_USERNAME/LFLB_PASSWORD.")

    session.get(f"{BASE_URL}/admin/signout", timeout=30, allow_redirects=True)
    login_page = session.get(f"{BASE_URL}/admin/login", timeout=30)
    login_page.raise_for_status()
    form = parse_login_form(login_page.text)
    action = urljoin(login_page.url, form["attrs"].get("action") or login_page.url)
    payload = fill_login_payload(dict(form["fields"]), username, password, tenant_code)

    response = session.post(
        action,
        headers={
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-language": "ja,en-US;q=0.9,en;q=0.8",
            "cache-control": "max-age=0",
            "content-type": "application/x-www-form-urlencoded",
            "origin": BASE_URL,
            "priority": "u=0, i",
            "referer": f"{BASE_URL}/admin/signout",
            "sec-ch-ua": '"Google Chrome";v="149", "Chromium";v="149", "Not)A;Brand";v="24"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "same-origin",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
        },
        data=payload,
        timeout=30,
        allow_redirects=True,
    )

    if debug:
        print(f"[debug] login form action={form['attrs'].get('action')!r}")
        print(f"[debug] login form method={form['attrs'].get('method')!r}")
        print(f"[debug] login form fields={sorted(form['fields'].keys())}")
        print(f"[debug] login payload keys={sorted(payload.keys())}")
        print(f"[debug] login status={response.status_code} url={response.url}")
        print(f"[debug] login response cookies={response.cookies.get_dict()}")
        print(f"[debug] session cookies={session.cookies.get_dict()}")
        print(f"[debug] login body head={response.text[:500]!r}")

    response.raise_for_status()
    if "JSESSIONID" not in session.cookies.get_dict() and "JSESSIONID" not in response.cookies.get_dict():
        raise RuntimeError("Login did not return a JSESSIONID cookie.")


def fetch_staff_page(session, staff_name, staff_code, work_date, page, rows, debug=False):
    payload = {
        "workStartTime": work_date.isoformat(),
        "workEndTime": work_date.isoformat(),
        "_search": "false",
        "nd": str(int(time.time() * 1000)),
        "rows": str(rows),
        "page": str(page),
        "sidx": "",
        "sord": "asc",
        "staffCode": staff_code,
        "staffName": staff_name,
        "clockOutPending": "",
    }
    response = session.post(
        f"{BASE_URL}/admin/api/workAttendanceRest/staff-list",
        headers={
            "accept": "application/json, text/javascript, */*; q=0.01",
            "content-type": "application/x-www-form-urlencoded; charset=UTF-8",
            "referer": f"{BASE_URL}/admin/workAttendance/staff-list",
            "x-requested-with": "XMLHttpRequest",
        },
        data=payload,
        timeout=30,
    )
    response.raise_for_status()
    if response.headers.get("content-type", "").lower().startswith("text/html"):
        raise RuntimeError(f"Expected JSON but received HTML for {staff_name} on {work_date}: {response.text[:300]!r}")
    if debug:
        print(f"[debug] query payload={payload}")
        print(f"[debug] staff-list status={response.status_code}")
        print(f"[debug] staff-list body head={response.text[:500]!r}")
    return response.json()


def iter_staff_records(session, staff_name, staff_code, work_date, rows, debug=False):
    page = 1
    while True:
        payload = fetch_staff_page(session, staff_name, staff_code, work_date, page, rows, debug=debug)
        if isinstance(payload, dict):
            page_rows = payload.get("rows")
            if page_rows is None:
                page_rows = payload.get("list", [])
        else:
            page_rows = []
        if debug and page_rows:
            first_row = page_rows[0]
            if isinstance(first_row, dict):
                print(f"[debug] row keys={sorted(first_row.keys())}")
                print(f"[debug] sample record={normalize_record(staff_name, staff_code, work_date, first_row)}")
        for row in page_rows:
            yield normalize_record(staff_name, staff_code, work_date, row)

        if isinstance(payload, dict):
            total_value = payload.get("total")
            total_pages = int(total_value or page)
        else:
            total_pages = page
        if page >= total_pages or not page_rows:
            break
        page += 1


def normalize_record(requested_name, requested_code, work_date, row):
    record = {
        "query_staff_name": requested_name,
        "query_staff_code": requested_code,
        "query_work_date": work_date.isoformat(),
    }

    if isinstance(row, dict):
        staff = row.get("staff") if isinstance(row.get("staff"), dict) else {}
        company = row.get("company") if isinstance(row.get("company"), dict) else {}
        work_company = row.get("workCompany") if isinstance(row.get("workCompany"), dict) else {}
        record.update(
            {
                "staffCode": first_value(row, "staffCode", "workStaffCode", "staffNo", "workStaffNo", fallback=first_value(staff, "staffCode", "staffNo", "workerNumber", fallback=requested_code)),
                "staffName": first_value(row, "staffName", "workStaffName", "name", fallback=first_value(staff, "staffName", "name", fallback=requested_name)),
                "companyName": first_nonempty(
                    [
                        deep_find(row, "companyName"),
                        deep_find(row, "workCompanyName"),
                        deep_find(company, "companyName"),
                        deep_find(company, "name"),
                        deep_find(work_company, "companyName"),
                        deep_find(work_company, "name"),
                    ],
                    "",
                ),
                "businessLevel": first_value(row, "businessLevel", "workBusinessLevel", "bussinessGrade", "levelName", "jobLevel", "position", "jobTitle", fallback=first_value(staff, "bussinessGrade", "businessLevel", fallback="")),
                "workStartTime": compact_time(row.get("workStartTime")),
                "realWorkStartTime": compact_time(row.get("realWorkStartTime")),
                "workEndTime": compact_time(row.get("workEndTime")),
                "realWorkEndTime": compact_time(row.get("realWorkEndTime")),
            }
        )
        return record

    return record


def compact_time(value):
    if isinstance(value, str) and len(value) >= 16:
        return value[:16]
    return value


def first_value(row, *keys, fallback=""):
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return fallback


def first_nonempty(values, fallback=""):
    for value in values:
        if value not in (None, ""):
            return value
    return fallback


def deep_find(value, target_key):
    if isinstance(value, dict):
        for key, item in value.items():
            if key == target_key and item not in (None, ""):
                return item
            found = deep_find(item, target_key)
            if found not in (None, ""):
                return found
    elif isinstance(value, list):
        for item in value:
            found = deep_find(item, target_key)
            if found not in (None, ""):
                return found
    return ""


def export_xlsx(records, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    columns = [
        "staffCode",
        "staffName",
        "companyName",
        "businessLevel",
        "workStartTime",
        "realWorkStartTime",
        "workEndTime",
        "realWorkEndTime",
    ]

    ws.append(columns)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    for record in sorted(records, key=lambda item: (item.get("staffName", ""), item.get("workStartTime", ""), item.get("staffCode", ""))):
        ws.append([safe_excel_value(record.get(column, "")) for column in columns])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_index, column_name in enumerate(columns, start=1):
        values = [column_name] + [str(record.get(column_name, "")) for record in records[:200]]
        width = min(max(len(value) for value in values) + 2, 60)
        ws.column_dimensions[get_column_letter(column_index)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def safe_excel_value(value):
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def main():
    args = parse_args()
    employees = load_employees(args.employees)
    dates = date_range(args.end_date, args.days)
    output_path = Path(args.output or f"outputs/attendance_{dates[0]}_to_{dates[-1]}.xlsx")

    session = make_session(args.tenant_code)
    login(session, args.username, args.password, args.tenant_code, debug=args.debug)

    records = []
    for work_date in dates:
        for employee in employees:
            name = employee["name"]
            code = employee["code"]
            records.extend(iter_staff_records(session, name, code, work_date, args.rows, debug=args.debug))

    export_xlsx(records, output_path)
    print(f"Exported {len(records)} rows to {output_path}")


if __name__ == "__main__":
    main()
